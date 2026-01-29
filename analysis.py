import subprocess
import pandas as pd
from datetime import datetime
import json
import os
import io

# === CONFIGURATION ===
ACCESS_PATH = "/home/bomar-ubu-1/migration-access/risk.mdb"  # WSL path to your .mdb file
OUTPUT_DIR = "access_analysis"

# Thematic subdirectories for organized output
SCHEMA_DIR = "01-schema"
POWERBI_DIR = "02-powerbi"
DATA_QUALITY_DIR = "03-data-quality"
RELATIONSHIPS_DIR = "04-relationships"
MIGRATION_DIR = "05-migration-planning"
ETL_DIR = "06-etl"
ANALYSIS_DIR = "07-analysis"


class AccessDatabaseAnalyzerWSL:
    def __init__(self, db_path):
        self.db_path = db_path
        self.report = {}
        
        # Verify mdbtools is installed
        try:
            subprocess.run(["mdb-tables", "--version"], capture_output=True, check=True)
        except FileNotFoundError:
            raise RuntimeError("mdbtools not installed. Run: sudo apt install mdbtools")
        
        # Verify file exists
        if not os.path.exists(db_path):
            raise FileNotFoundError(f"Database not found: {db_path}")
    
    def run_mdb_command(self, command, *args):
        """Run an mdbtools command and return output"""
        try:
            result = subprocess.run(
                [command, self.db_path, *args],
                capture_output=True,
                text=True,
                check=True
            )
            return result.stdout
        except subprocess.CalledProcessError as e:
            print(f"   Warning: {command} failed - {e.stderr}")
            return ""
    
    def analyze_all(self):
        """Run complete database analysis"""
        print("=" * 60)
        print("ACCESS DATABASE ANALYZER (WSL/Linux)")
        print("=" * 60)
        print(f"Database: {self.db_path}")
        print(f"Analysis started: {datetime.now()}\n")
        
        self.report["database_path"] = self.db_path
        self.report["analysis_date"] = str(datetime.now())
        
        self.get_tables()
        self.get_queries()
        self.get_relationships()
        self.analyze_table_details()
        self.analyze_data_quality()
        self.identify_potential_issues()

        # Enhanced analysis methods
        self.detect_primary_keys()
        self.analyze_indexes()
        self.analyze_powerbi_impact()
        self.infer_foreign_keys()
        self.validate_referential_integrity()
        self.analyze_dax_impact()
        self.detect_dead_columns()

        return self.report
    
    def get_tables(self):
        """Get all user tables"""
        print("Analyzing tables...")
        
        output = self.run_mdb_command("mdb-tables", "-1")
        tables = [t.strip() for t in output.strip().split("\n") if t.strip()]
        
        # Filter out system tables
        tables = [t for t in tables if not t.startswith("MSys")]
        
        self.report["tables"] = {
            "count": len(tables),
            "names": tables
        }
        print(f"   Found {len(tables)} tables\n")
        return tables
    
    def get_queries(self):
        """Get saved queries using mdb-queries"""
        print("Analyzing saved queries...")
        
        queries = []
        
        try:
            output = self.run_mdb_command("mdb-queries", "-L")
            if output.strip():
                for line in output.strip().split("\n"):
                    if line.strip():
                        queries.append({
                            "name": line.strip(),
                            "type": "QUERY",
                            "sql": None
                        })
                
                # Try to get SQL for each query
                for q in queries:
                    try:
                        sql_output = self.run_mdb_command("mdb-queries", q["name"])
                        q["sql"] = sql_output.strip()[:500]  # Limit length
                    except:
                        pass
        except Exception as e:
            print(f"   Warning: Could not read queries - {e}")
        
        self.report["queries"] = {
            "count": len(queries),
            "details": queries
        }
        print(f"   Found {len(queries)} saved queries\n")
        
        if queries:
            print("   WARNING: Review these queries - they may need recreation in PostgreSQL\n")
        
        return queries
    
    def get_relationships(self):
        """Get relationships using mdb-schema"""
        print("Analyzing relationships...")
        
        relationships = []
        
        try:
            # Get schema which includes relationship info
            output = self.run_mdb_command("mdb-schema", "--relationships")
            
            # Parse ALTER TABLE ... ADD CONSTRAINT ... FOREIGN KEY statements
            for line in output.split("\n"):
                if "FOREIGN KEY" in line.upper() and "REFERENCES" in line.upper():
                    # Basic parsing - may need refinement
                    relationships.append({
                        "definition": line.strip(),
                        "parsed": False
                    })
        except Exception as e:
            print(f"   Warning: Could not read relationships - {e}")
        
        self.report["relationships"] = {
            "count": len(relationships),
            "details": relationships
        }
        print(f"   Found {len(relationships)} relationships\n")
        return relationships
    
    def get_table_schema(self, table_name):
        """Get column info for a table using mdb-schema"""
        columns = []
        
        try:
            output = self.run_mdb_command("mdb-schema", "-T", table_name)
            
            # Parse CREATE TABLE statement
            in_create = False
            for line in output.split("\n"):
                line = line.strip()
                
                if line.upper().startswith("CREATE TABLE"):
                    in_create = True
                    continue
                
                if in_create and line.startswith(")"):
                    break
                
                if in_create and line and not line.startswith("("):
                    # Parse column definition
                    # Format: column_name TYPE(size) [NOT NULL]
                    parts = line.rstrip(",").split()
                    if parts:
                        col_name = parts[0].strip("[]\"")
                        col_type = parts[1] if len(parts) > 1 else "VARCHAR"
                        
                        # Extract size if present
                        size = None
                        if "(" in col_type:
                            type_parts = col_type.split("(")
                            col_type = type_parts[0]
                            size = type_parts[1].rstrip(")")
                            try:
                                size = int(size.split(",")[0])
                            except:
                                size = None
                        
                        nullable = "NOT NULL" not in line.upper()
                        
                        columns.append({
                            "name": col_name,
                            "pg_name": col_name.lower().replace(" ", "_").replace("-", "_"),
                            "type": col_type.upper(),
                            "size": size,
                            "decimal_digits": None,
                            "nullable": nullable,
                            "ordinal": len(columns) + 1
                        })
        except Exception as e:
            print(f"      Error getting schema: {e}")
        
        return columns
    
    def export_table_to_df(self, table_name):
        """Export a table to pandas DataFrame"""
        try:
            output = self.run_mdb_command("mdb-export", table_name)
            if output:
                return pd.read_csv(io.StringIO(output))
        except Exception as e:
            print(f"      Error exporting {table_name}: {e}")
        return pd.DataFrame()
    
    def analyze_table_details(self):
        """Detailed analysis of each table"""
        print("Analyzing table structures...")
        
        table_details = []
        
        for table in self.report["tables"]["names"]:
            print(f"   Processing: {table}")
            
            detail = {
                "name": table,
                "pg_name": table.lower().replace(" ", "_").replace("-", "_"),
                "columns": [],
                "row_count": 0,
                "primary_key": None
            }
            
            # Get columns
            detail["columns"] = self.get_table_schema(table)
            
            # Get row count by exporting and counting
            try:
                df = self.export_table_to_df(table)
                detail["row_count"] = len(df)
            except:
                detail["row_count"] = 0
            
            table_details.append(detail)
        
        self.report["table_details"] = table_details
        print()
        return table_details
    
    def analyze_data_quality(self):
        """Analyze data quality"""
        print("Analyzing data quality (this may take a while)...")
        
        quality_report = []
        
        for table in self.report["table_details"]:
            table_name = table["name"]
            print(f"   Profiling: {table_name}")
            
            table_quality = {
                "table": table_name,
                "columns": []
            }
            
            try:
                df = self.export_table_to_df(table_name)
                
                for col in table["columns"]:
                    col_name = col["name"]
                    col_quality = {
                        "column": col_name,
                        "null_count": 0,
                        "null_percent": 0,
                        "distinct_count": 0,
                        "sample_values": []
                    }
                    
                    if col_name in df.columns:
                        col_quality["null_count"] = int(df[col_name].isna().sum())
                        if len(df) > 0:
                            col_quality["null_percent"] = round(
                                (col_quality["null_count"] / len(df)) * 100, 2
                            )
                        col_quality["distinct_count"] = int(df[col_name].nunique())
                        col_quality["sample_values"] = [
                            str(v)[:50] for v in df[col_name].dropna().unique()[:5]
                        ]
                    
                    table_quality["columns"].append(col_quality)
            except Exception as e:
                print(f"      Error profiling: {e}")
            
            quality_report.append(table_quality)
        
        self.report["data_quality"] = quality_report
        print()
        return quality_report
    
    def identify_potential_issues(self):
        """Identify potential migration issues"""
        print("Identifying potential migration issues...")
        
        issues = []
        pg_reserved = ['user', 'order', 'table', 'group', 'select', 'where', 'index', 
                       'key', 'primary', 'foreign', 'check', 'default', 'constraint']
        
        for table in self.report["table_details"]:
            table_name = table["name"]
            pg_name = table["pg_name"]
            
            if table_name != pg_name:
                issues.append({
                    "type": "NAMING",
                    "severity": "MEDIUM",
                    "table": table_name,
                    "column": None,
                    "issue": f"Table name will change: '{table_name}' -> '{pg_name}'",
                    "action": "Update Power BI queries to use new name"
                })
            
            if pg_name.lower() in pg_reserved:
                issues.append({
                    "type": "RESERVED_WORD",
                    "severity": "HIGH",
                    "table": table_name,
                    "column": None,
                    "issue": f"'{pg_name}' is a PostgreSQL reserved word",
                    "action": "Rename table or use quoted identifiers"
                })
            
            for col in table["columns"]:
                col_name = col["name"]
                pg_col_name = col["pg_name"]
                
                if col_name != pg_col_name:
                    issues.append({
                        "type": "NAMING",
                        "severity": "MEDIUM",
                        "table": table_name,
                        "column": col_name,
                        "issue": f"Column name will change: '{col_name}' -> '{pg_col_name}'",
                        "action": "Update Power BI queries"
                    })
                
                if pg_col_name.lower() in pg_reserved:
                    issues.append({
                        "type": "RESERVED_WORD",
                        "severity": "HIGH",
                        "table": table_name,
                        "column": col_name,
                        "issue": f"'{pg_col_name}' is a PostgreSQL reserved word",
                        "action": "Rename column or use quoted identifiers"
                    })
                
                if col["type"] in ["LONGBINARY", "OLE"]:
                    issues.append({
                        "type": "DATA_TYPE",
                        "severity": "HIGH",
                        "table": table_name,
                        "column": col_name,
                        "issue": "OLE Object field - may contain embedded files/images",
                        "action": "Decide how to handle binary data"
                    })
            
            if not table.get("primary_key"):
                issues.append({
                    "type": "SCHEMA",
                    "severity": "MEDIUM",
                    "table": table_name,
                    "column": None,
                    "issue": "No primary key defined",
                    "action": "Consider adding a primary key in PostgreSQL"
                })
        
        if self.report["queries"]["count"] > 0:
            issues.append({
                "type": "QUERIES",
                "severity": "HIGH",
                "table": None,
                "column": None,
                "issue": f"{self.report['queries']['count']} saved queries found",
                "action": "Review and recreate as PostgreSQL views or Power BI queries"
            })
        
        self.report["potential_issues"] = {
            "count": len(issues),
            "by_severity": {
                "HIGH": len([i for i in issues if i["severity"] == "HIGH"]),
                "MEDIUM": len([i for i in issues if i["severity"] == "MEDIUM"]),
                "LOW": len([i for i in issues if i["severity"] == "LOW"])
            },
            "details": issues
        }
        
        print(f"   Found {len(issues)} potential issues")
        print(f"   - HIGH: {self.report['potential_issues']['by_severity']['HIGH']}")
        print(f"   - MEDIUM: {self.report['potential_issues']['by_severity']['MEDIUM']}")
        print(f"   - LOW: {self.report['potential_issues']['by_severity']['LOW']}\n")
        
        return issues

    def detect_primary_keys(self):
        """Detect primary keys using schema analysis and data patterns"""
        print("Detecting primary keys...")

        for table in self.report["table_details"]:
            table_name = table["name"]

            # Method 1: Parse mdb-schema for PRIMARY KEY keywords
            try:
                schema = self.run_mdb_command("mdb-schema", "-T", table_name)
                if "PRIMARY KEY" in schema.upper():
                    # Extract PK column names from schema
                    for line in schema.split("\n"):
                        if "PRIMARY KEY" in line.upper():
                            # Parse: PRIMARY KEY (`column_name`)
                            import re
                            pk_match = re.search(r'PRIMARY KEY\s*\(\s*[`"\[]?(\w+)[`"\]]?\s*\)', line, re.IGNORECASE)
                            if pk_match:
                                table["primary_key"] = pk_match.group(1)
                                table["primary_key_type"] = "defined"
                                continue
            except Exception as e:
                print(f"      Error parsing PK from schema: {e}")

            # Method 2: Check for ID column (likely auto-number)
            if not table.get("primary_key"):
                for col in table["columns"]:
                    if col["name"].upper() == "ID" and col["type"] in ["LONG INTEGER", "COUNTER"]:
                        table["primary_key"] = col["name"]
                        table["primary_key_type"] = "auto_number_id"
                        break

            # Method 3: Check for caceis_id (business key)
            if not table.get("primary_key"):
                for col in table["columns"]:
                    if col["name"].lower() == "caceis_id" and not col["nullable"]:
                        table["primary_key"] = col["name"]
                        table["primary_key_type"] = "business_key"
                        break

            # Method 4: Analyze data for uniqueness (check first column that's unique)
            if not table.get("primary_key"):
                try:
                    df = self.export_table_to_df(table_name)
                    if len(df) > 0:
                        for col in table["columns"]:
                            col_name = col["name"]
                            if col_name in df.columns:
                                # Check if column has all unique non-null values
                                non_null_count = df[col_name].notna().sum()
                                unique_count = df[col_name].nunique()
                                if non_null_count > 0 and non_null_count == unique_count == len(df):
                                    table["primary_key"] = col_name
                                    table["primary_key_type"] = "inferred_unique"
                                    break
                except:
                    pass

        pk_found = sum(1 for t in self.report["table_details"] if t.get("primary_key"))
        print(f"   Detected primary keys in {pk_found}/{len(self.report['table_details'])} tables\n")

    def analyze_indexes(self):
        """Get index information from Access database"""
        print("Analyzing indexes...")

        indexes = []

        for table in self.report["table_details"]:
            table_name = table["name"]

            try:
                # Try to get index info from schema
                schema = self.run_mdb_command("mdb-schema", "-T", table_name, "--indexes")

                # Parse CREATE INDEX statements
                for line in schema.split("\n"):
                    if "CREATE" in line.upper() and "INDEX" in line.upper():
                        import re
                        # Parse: CREATE [UNIQUE] INDEX index_name ON table_name (columns)
                        index_match = re.search(r'CREATE\s+(UNIQUE\s+)?INDEX\s+(\w+)\s+ON\s+\w+\s*\(([^)]+)\)', line, re.IGNORECASE)
                        if index_match:
                            indexes.append({
                                "table": table_name,
                                "index_name": index_match.group(2),
                                "columns": index_match.group(3).strip(),
                                "is_unique": bool(index_match.group(1)),
                                "is_primary_key": False,
                                "postgres_recommendation": "CREATE INDEX" if not index_match.group(1) else "CREATE UNIQUE INDEX"
                            })
            except Exception as e:
                print(f"      Error analyzing indexes for {table_name}: {e}")

        # Mark primary key indexes
        for idx in indexes:
            for table in self.report["table_details"]:
                if table["name"] == idx["table"] and table.get("primary_key"):
                    if table["primary_key"] in idx["columns"]:
                        idx["is_primary_key"] = True

        self.report["indexes"] = indexes
        print(f"   Found {len(indexes)} indexes\n")
        return indexes

    def analyze_powerbi_impact(self):
        """Analyze specific Power BI migration impacts"""
        print("Analyzing Power BI impact...")

        powerbi_impacts = []

        for table in self.report["table_details"]:
            impact = {
                "access_table_name": table["name"],
                "postgres_table_name": table["pg_name"],
                "name_will_change": "YES" if table["name"] != table["pg_name"] else "NO",
                "m_query_update_required": "YES" if table["name"] != table["pg_name"] else "NO",
                "column_count": len(table["columns"]),
                "columns_with_name_changes": 0,
                "columns_with_special_chars": 0,
                "complexity_score": 0.0,
                "migration_risk": "",
                "recommended_approach": "",
                "example_m_query_before": f'Source = Access.Database(File.Contents("{self.db_path}"), [Name="{table["name"]}"])',
                "example_m_query_after": f'Source = PostgreSQL.Database("server", "database")[{table["pg_name"]}]'
            }

            # Count column issues
            for col in table["columns"]:
                if col["name"] != col["pg_name"]:
                    impact["columns_with_name_changes"] += 1
                if any(c in col["name"] for c in [" ", "(", ")", "/", "%", "-"]):
                    impact["columns_with_special_chars"] += 1

            # Calculate complexity score (0-10 scale)
            impact["complexity_score"] = min(10, (
                (5 if impact["name_will_change"] == "YES" else 0) +
                (impact["columns_with_name_changes"] * 0.3) +
                (impact["columns_with_special_chars"] * 0.2)
            ))
            impact["complexity_score"] = round(impact["complexity_score"], 1)

            # Assign risk level
            if impact["complexity_score"] >= 7:
                impact["migration_risk"] = "HIGH"
                impact["recommended_approach"] = "VIEW - Use compatibility view initially"
            elif impact["complexity_score"] >= 3:
                impact["migration_risk"] = "MEDIUM"
                impact["recommended_approach"] = "DIRECT or VIEW - Test both approaches"
            else:
                impact["migration_risk"] = "LOW"
                impact["recommended_approach"] = "DIRECT - Minimal changes needed"

            powerbi_impacts.append(impact)

        self.report["powerbi_impact"] = powerbi_impacts
        print(f"   Analyzed {len(powerbi_impacts)} tables for Power BI impact\n")
        return powerbi_impacts

    def generate_naming_mapping(self):
        """Generate comprehensive before/after naming mapping"""
        print("Generating naming mappings...")

        # Table-level mapping
        table_mappings = []
        for table in self.report["table_details"]:
            table_mappings.append({
                "object_type": "TABLE",
                "access_name": table["name"],
                "postgres_name": table["pg_name"],
                "name_changed": "YES" if table["name"] != table["pg_name"] else "NO",
                "requires_quotes_in_postgres": "YES" if table["name"] != table["pg_name"] else "NO",
                "powerbi_search_pattern": f'[Name="{table["name"]}"]',
                "powerbi_replace_with": f'use PostgreSQL connector with table: {table["pg_name"]}'
            })

        # Column-level mapping
        column_mappings = []
        for table in self.report["table_details"]:
            for col in table["columns"]:
                if col["name"] != col["pg_name"]:
                    # Identify change reason
                    change_reason = []
                    if " " in col["name"]:
                        change_reason.append("spaces")
                    if "(" in col["name"] or ")" in col["name"]:
                        change_reason.append("parentheses")
                    if "/" in col["name"]:
                        change_reason.append("slashes")
                    if "%" in col["name"]:
                        change_reason.append("percent")
                    if "-" in col["name"]:
                        change_reason.append("hyphens")

                    column_mappings.append({
                        "object_type": "COLUMN",
                        "table_access": table["name"],
                        "table_postgres": table["pg_name"],
                        "column_access": col["name"],
                        "column_postgres": col["pg_name"],
                        "change_reason": ", ".join(change_reason) if change_reason else "other",
                        "powerbi_search_pattern": f'[{col["name"]}]',
                        "powerbi_replace_with": col["pg_name"],
                        "dax_search_pattern": f'{table["name"]}[{col["name"]}]',
                        "dax_replace_with": f'{table["pg_name"]}[{col["pg_name"]}]'
                    })

        print(f"   Generated {len(table_mappings)} table mappings")
        print(f"   Generated {len(column_mappings)} column mappings\n")

        return table_mappings, column_mappings

    def generate_compatibility_views(self, filepath):
        """Generate PostgreSQL views that preserve Access naming"""
        print("   Generating compatibility views...")

        views_created = 0

        with open(filepath, "w", encoding="utf-8") as f:
            f.write("-- PostgreSQL Compatibility Views\n")
            f.write("-- These views preserve Access table/column names for Power BI compatibility\n")
            f.write(f"-- Generated: {datetime.now()}\n")
            f.write(f"-- Source: {self.db_path}\n\n")
            f.write("-- USAGE: Point Power BI to these views initially, then gradually migrate to base tables\n\n")

            for table in self.report["table_details"]:
                # Create view if table or any column name changed
                has_changes = (table["name"] != table["pg_name"]) or any(
                    col["name"] != col["pg_name"] for col in table["columns"]
                )

                if has_changes:
                    view_name = f"{table['pg_name']}_compat_view"

                    f.write(f"-- Compatibility view for Access table: {table['name']}\n")
                    f.write(f"DROP VIEW IF EXISTS \"{view_name}\" CASCADE;\n")
                    f.write(f"CREATE OR REPLACE VIEW \"{view_name}\" AS\n")
                    f.write("SELECT\n")

                    # Map columns back to original names if needed
                    col_mappings = []
                    for col in table["columns"]:
                        if col["name"] != col["pg_name"]:
                            # Rename back to Access name
                            col_mappings.append(f'    "{col["pg_name"]}" AS "{col["name"]}"')
                        else:
                            col_mappings.append(f'    "{col["pg_name"]}"')

                    f.write(",\n".join(col_mappings))
                    f.write(f'\nFROM "{table["pg_name"]}";\n\n')

                    # Add comment
                    f.write(f"COMMENT ON VIEW \"{view_name}\" IS 'Compatibility view preserving Access names for table: {table['name']}';\n\n")

                    views_created += 1

        print(f"      Created {views_created} compatibility views")
        return views_created

    def infer_foreign_keys(self):
        """Infer foreign key relationships from naming patterns and data"""
        print("Inferring foreign key relationships...")

        inferred_fks = []

        # Get LIST FUNDS caceis_id values (the main reference table)
        list_funds_ids = set()
        try:
            df = self.export_table_to_df("LIST FUNDS")
            if "caceis_id" in df.columns:
                list_funds_ids = set(df["caceis_id"].dropna().unique())
        except:
            pass

        for table in self.report["table_details"]:
            table_name = table["name"]

            # Skip the reference table itself
            if table_name == "LIST FUNDS":
                continue

            try:
                df = self.export_table_to_df(table_name)

                for col in table["columns"]:
                    col_name = col["name"]

                    # Pattern 1: caceis_id references LIST FUNDS
                    if col_name.lower() == "caceis_id" and col_name in df.columns:
                        # Check if values match LIST FUNDS
                        if list_funds_ids:
                            col_values = set(df[col_name].dropna().unique())
                            if col_values and col_values.issubset(list_funds_ids):
                                confidence = "HIGH"
                            elif col_values and len(col_values.intersection(list_funds_ids)) > 0:
                                confidence = "MEDIUM"
                            else:
                                confidence = "LOW"

                            inferred_fks.append({
                                "from_table": table_name,
                                "from_column": col_name,
                                "to_table": "LIST FUNDS",
                                "to_column": "caceis_id",
                                "confidence": confidence,
                                "pattern": "exact_name_match",
                                "postgres_fk_sql": f'ALTER TABLE "{table["pg_name"]}" ADD CONSTRAINT "fk_{table["pg_name"]}_caceis_id" FOREIGN KEY ("{col["pg_name"]}") REFERENCES "list_funds" ("caceis_id");'
                            })

                    # Pattern 2: _id or _code suffix columns
                    elif col_name.lower().endswith("_id") or col_name.lower().endswith("_code"):
                        # Try to find referenced table
                        potential_ref = col_name.lower().replace("_id", "").replace("_code", "")
                        # Look for matching table
                        for ref_table in self.report["table_details"]:
                            if potential_ref in ref_table["name"].lower():
                                inferred_fks.append({
                                    "from_table": table_name,
                                    "from_column": col_name,
                                    "to_table": ref_table["name"],
                                    "to_column": "ID",  # Assumption
                                    "confidence": "LOW",
                                    "pattern": "naming_convention",
                                    "postgres_fk_sql": f"-- Verify and create: FK from {table['pg_name']}.{col['pg_name']} to {ref_table['pg_name']}.id"
                                })
                                break

            except Exception as e:
                print(f"      Error analyzing {table_name}: {e}")

        self.report["inferred_foreign_keys"] = inferred_fks
        print(f"   Inferred {len(inferred_fks)} potential foreign key relationships\n")
        return inferred_fks

    def generate_connection_docs(self, output_dir):
        """Generate PostgreSQL connection documentation for Power BI"""
        filepath = f"{output_dir}/powerbi_connection_guide.md"

        with open(filepath, "w", encoding="utf-8") as f:
            f.write("# Power BI PostgreSQL Connection Guide\n\n")
            f.write(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\\n\\n")

            f.write("## Connection Information\n\n")
            f.write("```\n")
            f.write("Server: <your-postgres-server>\n")
            f.write("Database: <your-database-name>\n")
            f.write("Port: 5432 (default)\n")
            f.write("```\n\n")

            f.write("## Step-by-Step Connection in Power BI Desktop\n\n")
            f.write("1. Open Power BI Desktop\n")
            f.write("2. Click **Get Data** → **More**\n")
            f.write("3. Search for **PostgreSQL database**\n")
            f.write("4. Click **Connect**\n")
            f.write("5. Enter:\n")
            f.write("   - **Server:** your-postgres-server\n")
            f.write("   - **Database:** your-database-name\n")
            f.write("6. Choose **Import** mode (recommended) or **DirectQuery**\n")
            f.write("7. Select tables to import\n\n")

            f.write("## M Query Migration Examples\n\n")

            # Show 3 examples
            examples = self.report["table_details"][:3]

            f.write("### Before (Access Database)\n\n")
            f.write("```powerquery\n")
            for table in examples:
                f.write(f'Source = Access.Database(\n')
                f.write(f'    File.Contents("{self.db_path}"),\n')
                f.write(f'    [Name="{table["name"]}"]\n')
                f.write(f')\n\n')
            f.write("```\n\n")

            f.write("### After (PostgreSQL - Direct Connection)\n\n")
            f.write("```powerquery\n")
            for table in examples:
                f.write(f'Source = PostgreSQL.Database(\n')
                f.write(f'    "your-server",\n')
                f.write(f'    "your-database"\n')
                f.write(f')[{table["pg_name"]}]\n\n')
            f.write("```\n\n")

            f.write("### After (PostgreSQL - Using Compatibility Views)\n\n")
            f.write("```powerquery\n")
            for table in examples:
                if table["name"] != table["pg_name"]:
                    f.write(f'Source = PostgreSQL.Database(\n')
                    f.write(f'    "your-server",\n')
                    f.write(f'    "your-database"\n')
                    f.write(f')[{table["pg_name"]}_compat_view]\n\n')
            f.write("```\n\n")

            f.write("## Important Notes\n\n")
            f.write("- **Table names:** All converted to lowercase with underscores\n")
            f.write("- **Column names:** Spaces replaced with underscores\n")
            f.write("- **Data types:** See data type mapping document\n")
            f.write("- **Compatibility views:** Use these for gradual migration\n\n")

            f.write("## Troubleshooting\n\n")
            f.write("### Connection Issues\n")
            f.write("- Verify PostgreSQL server is accessible\n")
            f.write("- Check firewall settings\n")
            f.write("- Verify username/password\n\n")

            f.write("### Query Performance\n")
            f.write("- Use Import mode for best performance\n")
            f.write("- DirectQuery recommended only for very large datasets\n")
            f.write("- Ensure indexes are created on join columns\n\n")

    def generate_migration_checklist_md(self, filepath):
        """Generate step-by-step migration checklist"""
        with open(filepath, "w", encoding="utf-8") as f:
            f.write("# Access to PostgreSQL Migration Checklist\n\n")
            f.write(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\\n\\n")
            f.write(f"**Source Database:** {self.db_path}\\n")
            f.write(f"**Total Tables:** {self.report['tables']['count']}\\n")
            f.write(f"**Total Rows:** {sum(t['row_count'] for t in self.report['table_details'] if isinstance(t['row_count'], int)):,}\\n\\n")

            f.write("## Phase 1: Pre-Migration (1-2 weeks before)\n\n")
            f.write("- [ ] Run complete analysis.py report\n")
            f.write("- [ ] Review all generated Excel files and documentation\n")
            f.write("- [ ] Document all 23 Power BI semantic models\n")
            f.write("- [ ] Export all Power BI M queries for reference\n")
            f.write("- [ ] Set up PostgreSQL test environment\n")
            f.write("- [ ] Install PostgreSQL ODBC driver on all Power BI machines\n")
            f.write("- [ ] Create test data subset\n")
            f.write("- [ ] Schedule stakeholder approval meeting\n")
            f.write("- [ ] Identify migration test team\n\n")

            f.write("## Phase 2: Schema Migration\n\n")
            f.write("- [ ] Create PostgreSQL database\n")
            f.write("- [ ] Run postgresql_schema.sql to create tables\n")
            f.write(f"- [ ] Add {sum(1 for t in self.report['table_details'] if not t.get('primary_key'))} missing primary keys\n")
            f.write(f"- [ ] Create {len(self.report.get('indexes', []))} indexes\n")
            f.write(f"- [ ] Add {len(self.report.get('inferred_foreign_keys', []))} foreign keys (review inferred_foreign_keys.xlsx)\n")
            f.write("- [ ] Run postgresql_compatibility_views.sql to create compatibility views\n")
            f.write("- [ ] Verify all tables created successfully\n")
            f.write("- [ ] Run GRANT statements for Power BI service account\n\n")

            f.write("## Phase 3: Data Migration\n\n")
            f.write(f"- [ ] Export data from Access using mdb-export for all {self.report['tables']['count']} tables\n")
            f.write("- [ ] Transform data (handle dates, booleans, nulls)\n")
            f.write("- [ ] Import to PostgreSQL using COPY command\n")
            f.write("- [ ] Run data_validation_queries.sql to verify data\n")
            f.write("- [ ] Compare row counts (Access vs PostgreSQL)\n")
            f.write("- [ ] Verify data types and null values\n")
            f.write("- [ ] Check for encoding issues\n\n")

            f.write("## Phase 4: Power BI Migration (Per Semantic Model)\n\n")
            f.write("### For each of 23 semantic models:\n\n")
            f.write("- [ ] 1. Create backup copy of .pbix file\n")
            f.write("- [ ] 2. Open in Power BI Desktop\n")
            f.write("- [ ] 3. Transform → Data source settings → Change source\n")
            f.write("- [ ] 4. Update to PostgreSQL connection\n")
            f.write("- [ ] 5. Update M queries using naming_mapping_tables.xlsx\n")
            f.write("- [ ] 6. Refresh data model (check for errors)\n")
            f.write("- [ ] 7. Verify all tables loaded successfully\n")
            f.write("- [ ] 8. Test all DAX measures (check dax_impact_analysis.xlsx)\n")
            f.write("- [ ] 9. Test all relationships\n")
            f.write("- [ ] 10. Test all reports/visuals\n")
            f.write("- [ ] 11. Compare results with Access version\n")
            f.write("- [ ] 12. User acceptance testing\n")
            f.write("- [ ] 13. Publish to Power BI Service\n")
            f.write("- [ ] 14. Test scheduled refresh\n\n")

            f.write("## Phase 5: Validation\n\n")
            f.write("- [ ] Run referential_integrity_issues.xlsx checks\n")
            f.write("- [ ] Compare Access vs PostgreSQL query results\n")
            f.write("- [ ] Performance testing (query response times)\n")
            f.write("- [ ] Verify scheduled refreshes work\n")
            f.write("- [ ] Check all dashboards render correctly\n")
            f.write("- [ ] Monitor for errors for 1 week\n\n")

            f.write("## Phase 6: Cutover\n\n")
            f.write("- [ ] Final communication to all users\n")
            f.write("- [ ] Switch all production semantic models\n")
            f.write("- [ ] Monitor Power BI Service for issues\n")
            f.write("- [ ] Keep Access database as backup for 2 weeks\n")
            f.write("- [ ] Document lessons learned\n")
            f.write("- [ ] Remove compatibility views (after 3 months, optional)\n")
            f.write("- [ ] Decommission Access database (after 6 months)\n\n")

            # Add table-specific notes
            f.write("## Table-Specific Migration Notes\n\n")

            high_risk_tables = [t for t in self.report.get("powerbi_impact", []) if t.get("migration_risk") == "HIGH"]
            if high_risk_tables:
                f.write("### High-Risk Tables (Require Extra Testing)\n\n")
                for table_impact in high_risk_tables[:10]:  # Top 10
                    f.write(f"- **{table_impact['access_table_name']}**\n")
                    f.write(f"  - PostgreSQL name: `{table_impact['postgres_table_name']}`\n")
                    f.write(f"  - Columns changing: {table_impact['columns_with_name_changes']}\n")
                    f.write(f"  - Complexity: {table_impact['complexity_score']}/10\n")
                    f.write(f"  - Approach: {table_impact['recommended_approach']}\n\n")

    def generate_data_validation_queries(self, filepath):
        """Generate SQL queries for Access vs PostgreSQL data validation"""
        print("   Generating data validation queries...")

        with open(filepath, "w", encoding="utf-8") as f:
            f.write("-- Data Validation Queries for PostgreSQL\n")
            f.write("-- Run these after data migration to verify integrity\n")
            f.write(f"-- Generated: {datetime.now()}\n\n")

            f.write("-- =====================================\n")
            f.write("-- SECTION 1: Row Count Validation\n")
            f.write("-- =====================================\n\n")

            for table in self.report["table_details"]:
                pg_name = table["pg_name"]
                expected_rows = table["row_count"]

                f.write(f"-- Validate row count for: {table['name']}\n")
                f.write(f"SELECT\n")
                f.write(f"    '{pg_name}' as table_name,\n")
                f.write(f"    COUNT(*) as actual_rows,\n")
                f.write(f"    {expected_rows} as expected_rows,\n")
                f.write(f"    CASE\n")
                f.write(f"        WHEN COUNT(*) = {expected_rows} THEN 'PASS'\n")
                f.write(f"        ELSE 'FAIL - Row count mismatch'\n")
                f.write(f"    END as status\n")
                f.write(f'FROM "{pg_name}";\n\n')

            f.write("\n-- =====================================\n")
            f.write("-- SECTION 2: NOT NULL Constraint Validation\n")
            f.write("-- =====================================\n\n")

            for table in self.report["table_details"]:
                pg_name = table["pg_name"]
                for col in table["columns"]:
                    if not col["nullable"]:
                        f.write(f"-- Check for NULL violations: {pg_name}.{col['pg_name']}\n")
                        f.write(f"SELECT '{pg_name}.{col['pg_name']}' as column_name, COUNT(*) as null_violations\n")
                        f.write(f'FROM "{pg_name}"\n')
                        f.write(f'WHERE "{col["pg_name"]}" IS NULL;\n')
                        f.write(f"-- Expected: 0 violations\n\n")

            f.write("\n-- =====================================\n")
            f.write("-- SECTION 3: Primary Key Uniqueness\n")
            f.write("-- =====================================\n\n")

            for table in self.report["table_details"]:
                if table.get("primary_key"):
                    pg_name = table["pg_name"]
                    pk_col = next((c["pg_name"] for c in table["columns"] if c["name"] == table["primary_key"]), None)
                    if pk_col:
                        f.write(f"-- Check PK uniqueness: {pg_name}.{pk_col}\n")
                        f.write(f"SELECT\n")
                        f.write(f"    '{pg_name}.{pk_col}' as pk_column,\n")
                        f.write(f'    "{pk_col}",\n')
                        f.write(f"    COUNT(*) as duplicate_count\n")
                        f.write(f'FROM "{pg_name}"\n')
                        f.write(f'GROUP BY "{pk_col}"\n')
                        f.write(f"HAVING COUNT(*) > 1;\n")
                        f.write(f"-- Expected: No rows (all PKs should be unique)\n\n")

            f.write("\n-- =====================================\n")
            f.write("-- SECTION 4: Data Type Validation\n")
            f.write("-- =====================================\n\n")

            f.write("-- Check for invalid date values\n")
            for table in self.report["table_details"]:
                pg_name = table["pg_name"]
                date_cols = [c for c in table["columns"] if c["type"] == "DATETIME"]
                for col in date_cols:
                    f.write(f"SELECT '{pg_name}.{col['pg_name']}' as column_name, MIN(\"{col['pg_name']}\") as min_date, MAX(\"{col['pg_name']}\") as max_date FROM \"{pg_name}\";\n")

            f.write("\n-- =====================================\n")
            f.write("-- SECTION 5: Summary Validation Report\n")
            f.write("-- =====================================\n\n")

            f.write("-- Generate summary of all tables\n")
            f.write("SELECT\n")
            f.write("    schemaname,\n")
            f.write("    tablename,\n")
            f.write("    pg_size_pretty(pg_total_relation_size(schemaname||'.'||tablename)) as size\n")
            f.write("FROM pg_tables\n")
            f.write("WHERE schemaname = 'public'\n")
            f.write("ORDER BY pg_total_relation_size(schemaname||'.'||tablename) DESC;\n")

        print(f"      Generated validation queries for {len(self.report['table_details'])} tables")

    def validate_referential_integrity(self):
        """Find orphan records and FK violations before migration"""
        print("Validating referential integrity...")

        integrity_issues = []

        # Get LIST FUNDS caceis_id values (the reference)
        try:
            list_funds_df = self.export_table_to_df("LIST FUNDS")
            if "caceis_id" in list_funds_df.columns:
                valid_caceis_ids = set(list_funds_df["caceis_id"].dropna().unique())

                # Tables that should reference LIST FUNDS
                referencing_tables = [
                    "Comments Dashboard", "Comments RnC", "Manual Input",
                    "Manual Input Date", "Param Commitment", "Statique dashboard",
                    "List_Mapping ptf", "List_Ptf_Limit"
                ]

                for ref_table in referencing_tables:
                    try:
                        df = self.export_table_to_df(ref_table)
                        if "caceis_id" in df.columns:
                            # Find orphan records
                            orphans = df[~df["caceis_id"].isin(valid_caceis_ids) & df["caceis_id"].notna()]
                            if len(orphans) > 0:
                                integrity_issues.append({
                                    "type": "ORPHAN_RECORD",
                                    "severity": "HIGH",
                                    "parent_table": "LIST FUNDS",
                                    "parent_column": "caceis_id",
                                    "child_table": ref_table,
                                    "child_column": "caceis_id",
                                    "orphan_count": len(orphans),
                                    "orphan_values": ", ".join(str(v) for v in orphans["caceis_id"].unique()[:5]),
                                    "action": "Clean up orphan records or add missing entries to LIST FUNDS before migration",
                                    "postgres_fk_constraint": f"FK from {ref_table} to LIST FUNDS will fail with {len(orphans)} orphans"
                                })
                    except Exception as e:
                        print(f"      Error checking {ref_table}: {e}")
        except Exception as e:
            print(f"      Error loading LIST FUNDS: {e}")

        self.report["referential_integrity_issues"] = integrity_issues
        print(f"   Found {len(integrity_issues)} referential integrity issues\n")
        return integrity_issues

    def analyze_dax_impact(self):
        """Identify DAX measures that will break due to column name changes"""
        print("Analyzing DAX impact...")

        dax_impacts = []

        for table in self.report["table_details"]:
            for col in table["columns"]:
                if col["name"] != col["pg_name"]:
                    # This column rename will break DAX
                    impact_level = "HIGH"
                    if " " in col["name"] or "(" in col["name"] or ")" in col["name"] or "/" in col["name"]:
                        impact_level = "CRITICAL"

                    dax_impacts.append({
                        "table_access": table["name"],
                        "table_postgres": table["pg_name"],
                        "column_access": col["name"],
                        "column_postgres": col["pg_name"],
                        "impact_level": impact_level,
                        "dax_search_pattern": f'{table["name"]}[{col["name"]}]',
                        "dax_replace_with": f'{table["pg_name"]}[{col["pg_name"]}]',
                        "alternate_search_1": f'[{col["name"]}]',
                        "alternate_search_2": f'"{col["name"]}"',
                        "notes": "Search ALL DAX measures, calculated columns, and calculated tables for this column reference"
                    })

        self.report["dax_impact"] = dax_impacts
        print(f"   Identified {len(dax_impacts)} column changes that will impact DAX\n")
        return dax_impacts

    def generate_etl_scripts(self, output_dir):
        """Generate data migration scripts"""
        print("   Generating ETL migration scripts...")

        # CSV files will be in the same ETL directory
        csv_dir = f"{output_dir}/csv"

        # 1. Export script (Bash - from Access to CSV)
        export_script = f"{output_dir}/01_export_from_access.sh"
        with open(export_script, "w", encoding="utf-8") as f:
            f.write("#!/bin/bash\n")
            f.write("# Export all tables from Access to CSV\n")
            f.write(f"# Generated: {datetime.now()}\n\n")

            f.write("# Create CSV output directory\n")
            f.write(f'mkdir -p "{csv_dir}"\n\n')

            f.write("echo 'Starting Access database export...'\n\n")

            for table in self.report["table_details"]:
                f.write(f"# Export: {table['name']}\n")
                f.write(f'echo "Exporting {table["name"]}..."\n')
                f.write(f'mdb-export "{self.db_path}" "{table["name"]}" > "{csv_dir}/{table["pg_name"]}.csv"\n')
                f.write(f'echo "  -> {table["row_count"]} rows"\n\n')

            f.write('echo "Export complete!"\n')

        # Make script executable
        import os
        os.chmod(export_script, 0o755)

        # 2. PostgreSQL COPY script (SQL - from CSV to PostgreSQL)
        import_script = f"{output_dir}/02_import_to_postgres.sql"
        with open(import_script, "w", encoding="utf-8") as f:
            f.write("-- Import CSV files to PostgreSQL\n")
            f.write(f"-- Generated: {datetime.now()}\n")
            f.write("-- Run this script as PostgreSQL superuser\n\n")

            f.write("-- Disable triggers during import for performance\n")
            f.write("SET session_replication_role = 'replica';\n\n")

            for table in self.report["table_details"]:
                f.write(f"-- Import: {table['name']} -> {table['pg_name']}\n")
                f.write(f"\\echo 'Importing {table['pg_name']}...'\n")
                f.write(f"\\COPY \"{table['pg_name']}\" FROM '{csv_dir}/{table['pg_name']}.csv' WITH (FORMAT csv, HEADER true, ENCODING 'UTF8', NULL '');\n\n")

            f.write("-- Re-enable triggers\n")
            f.write("SET session_replication_role = 'origin';\n\n")

            f.write("-- Update sequences for tables with auto-increment IDs\n")
            for table in self.report["table_details"]:
                if table.get("primary_key") and table.get("primary_key_type") == "auto_number_id":
                    pk_col = next((c["pg_name"] for c in table["columns"] if c["name"] == table["primary_key"]), None)
                    if pk_col:
                        f.write(f"SELECT setval(pg_get_serial_sequence('\"{table['pg_name']}\"', '{pk_col}'), COALESCE((SELECT MAX({pk_col}) FROM \"{table['pg_name']}\"), 1));\n")

            f.write("\n\\echo 'Import complete!'\n")

        # 3. Python transformation script (for data cleaning)
        transform_script = f"{output_dir}/03_transform_data.py"
        with open(transform_script, "w", encoding="utf-8") as f:
            f.write("#!/usr/bin/env python3\n")
            f.write('"""Data transformation script for Access to PostgreSQL migration"""\n')
            f.write("# This script handles data type conversions and cleaning\n")
            f.write(f"# Generated: {datetime.now()}\n\n")

            f.write("import pandas as pd\n")
            f.write("import os\n")
            f.write("from datetime import datetime\n\n")

            f.write(f'CSV_DIR = "{csv_dir}"\n')
            f.write(f'OUTPUT_DIR = "{csv_dir}_transformed"\n\n')

            f.write("os.makedirs(OUTPUT_DIR, exist_ok=True)\n\n")

            f.write("def transform_table(filename, transformations):\n")
            f.write("    df = pd.read_csv(os.path.join(CSV_DIR, filename))\n")
            f.write("    # Apply transformations\n")
            f.write("    for col, transform in transformations.items():\n")
            f.write("        if col in df.columns:\n")
            f.write("            if transform == 'datetime':\n")
            f.write("                df[col] = pd.to_datetime(df[col], errors='coerce')\n")
            f.write("            elif transform == 'boolean':\n")
            f.write("                df[col] = df[col].map({'Yes': True, 'No': False, '1': True, '0': False})\n")
            f.write("    df.to_csv(os.path.join(OUTPUT_DIR, filename), index=False)\n")
            f.write("    print(f'Transformed {filename}')\n\n")

            f.write("# Table-specific transformations\n")
            f.write("# Add your transformations here\n\n")

            f.write("print('Transformation complete!')\n")

        os.chmod(transform_script, 0o755)

        print(f"      Generated 3 ETL scripts: export, import, transform")

    def detect_dead_columns(self):
        """Find columns that are always null or have only one distinct value"""
        print("Detecting dead/unused columns...")

        dead_columns = []

        for table in self.report["table_details"]:
            table_name = table["name"]

            try:
                df = self.export_table_to_df(table_name)

                if len(df) == 0:
                    continue

                for col in table["columns"]:
                    col_name = col["name"]
                    if col_name in df.columns:
                        null_count = df[col_name].isna().sum()
                        null_pct = (null_count / len(df)) * 100 if len(df) > 0 else 0
                        distinct_count = df[col_name].nunique()

                        # Always null
                        if null_pct == 100:
                            dead_columns.append({
                                "table": table_name,
                                "column": col_name,
                                "issue": "ALWAYS_NULL",
                                "recommendation": "DISCARD - Column never used",
                                "null_percent": 100.0,
                                "distinct_count": 0,
                                "sample_value": None
                            })

                        # Only one value (and not a small lookup table)
                        elif distinct_count == 1 and len(df) > 10:
                            sample_val = str(df[col_name].dropna().iloc[0]) if len(df[col_name].dropna()) > 0 else None
                            dead_columns.append({
                                "table": table_name,
                                "column": col_name,
                                "issue": "SINGLE_VALUE",
                                "recommendation": "REVIEW - May be deprecated or constant",
                                "null_percent": null_pct,
                                "distinct_count": 1,
                                "sample_value": sample_val[:50] if sample_val else None
                            })

                        # Mostly null (> 95%)
                        elif null_pct > 95 and len(df) > 10:
                            dead_columns.append({
                                "table": table_name,
                                "column": col_name,
                                "issue": "MOSTLY_NULL",
                                "recommendation": "REVIEW - Rarely used",
                                "null_percent": round(null_pct, 2),
                                "distinct_count": distinct_count,
                                "sample_value": None
                            })

            except Exception as e:
                print(f"      Error analyzing {table_name}: {e}")

        self.report["dead_columns"] = dead_columns
        print(f"   Found {len(dead_columns)} potentially dead/unused columns\n")
        return dead_columns

    def generate_output_readme(self, output_dir):
        """Generate README.md to explain the output directory structure"""
        filepath = f"{output_dir}/README.md"

        with open(filepath, "w", encoding="utf-8") as f:
            f.write("# Access to PostgreSQL Migration Analysis\n\n")
            f.write(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\\n")
            f.write(f"**Source Database:** `{self.db_path}`\\n\\n")

            f.write("## Directory Structure\n\n")
            f.write("The analysis files are organized by migration theme:\n\n")

            f.write("### 📐 01-schema/\n")
            f.write("**Database structure files**\n")
            f.write("- `postgresql_schema.sql` - PostgreSQL CREATE TABLE statements\n")
            f.write("- `postgresql_compatibility_views.sql` - Views that preserve Access naming\n\n")

            f.write("### 📊 02-powerbi/\n")
            f.write("**Power BI migration guidance**\n")
            f.write("- `powerbi_connection_guide.md` - Step-by-step connection instructions\n")
            f.write("- `powerbi_impact_analysis.xlsx` - Impact of naming changes on reports\n")
            f.write("- `dax_impact_analysis.xlsx` - DAX formulas that need updates\n")
            f.write("- `naming_mapping_tables.xlsx` - Table name mappings (Access → PostgreSQL)\n")
            f.write("- `naming_mapping_columns.xlsx` - Column name mappings (Access → PostgreSQL)\n\n")

            f.write("### ✅ 03-data-quality/\n")
            f.write("**Data quality assessment and validation**\n")
            f.write("- `data_quality.xlsx` - Null counts, distinct values, sample data\n")
            f.write("- `issues.xlsx` - Detected migration issues and warnings\n")
            f.write("- `dead_columns_analysis.xlsx` - Unused or deprecated columns\n")
            f.write("- `referential_integrity_issues.xlsx` - Foreign key violations\n")
            f.write("- `data_validation_queries.sql` - Post-migration validation queries\n\n")

            f.write("### 🔗 04-relationships/\n")
            f.write("**Database relationships and keys**\n")
            f.write("- `relationships.xlsx` - Detected Access relationships\n")
            f.write("- `inferred_foreign_keys.xlsx` - AI-inferred foreign key relationships\n")
            f.write("- `indexes.xlsx` - Index definitions for PostgreSQL\n\n")

            f.write("### 📋 05-migration-planning/\n")
            f.write("**Migration planning and execution tracking**\n")
            f.write("- `MIGRATION_SUMMARY.md` - Executive summary\n")
            f.write("- `migration_checklist.md` - Step-by-step migration checklist\n")
            f.write("- `migration_review_checklist.xlsx` - Interactive stakeholder review form\n")
            f.write("- `full_analysis.json` - Complete raw analysis data\n\n")

            f.write("### 🔄 06-etl/\n")
            f.write("**ETL scripts and data export**\n")
            f.write("- `01_export_from_access.sh` - Bash script to export Access tables to CSV\n")
            f.write("- `02_import_to_postgres.sql` - SQL script to import CSV to PostgreSQL\n")
            f.write("- `03_transform_data.py` - Python script for data transformations\n")
            f.write("- `queries.xlsx` - Access saved queries (if any)\n")
            f.write("- `csv/` - Exported CSV files (created when running export script)\n\n")

            f.write("### 📈 07-analysis/\n")
            f.write("**Core database analysis tables**\n")
            f.write("- `tables_summary.xlsx` - Summary of all tables\n")
            f.write("- `columns_detail.xlsx` - Detailed column information\n\n")

            f.write("## Quick Start Guide\n\n")
            f.write("### For Database Administrators\n")
            f.write("1. Review `05-migration-planning/MIGRATION_SUMMARY.md`\n")
            f.write("2. Check `03-data-quality/issues.xlsx` for critical issues\n")
            f.write("3. Execute `01-schema/postgresql_schema.sql` to create tables\n")
            f.write("4. Run `06-etl/01_export_from_access.sh` to export data\n")
            f.write("5. Execute `06-etl/02_import_to_postgres.sql` to import data\n\n")

            f.write("### For Power BI Developers\n")
            f.write("1. Read `02-powerbi/powerbi_connection_guide.md`\n")
            f.write("2. Review `02-powerbi/powerbi_impact_analysis.xlsx`\n")
            f.write("3. Use `02-powerbi/naming_mapping_*.xlsx` to update queries\n")
            f.write("4. Check `02-powerbi/dax_impact_analysis.xlsx` for DAX updates\n\n")

            f.write("### For Project Managers\n")
            f.write("1. Start with `05-migration-planning/MIGRATION_SUMMARY.md`\n")
            f.write("2. Distribute `05-migration-planning/migration_review_checklist.xlsx` to stakeholders\n")
            f.write("3. Follow `05-migration-planning/migration_checklist.md` for execution plan\n\n")

            f.write("## File Counts\n\n")
            f.write(f"- **Total Files:** ~26 files\n")
            f.write(f"- **Excel Files:** ~15 files (.xlsx)\n")
            f.write(f"- **SQL Scripts:** 4 files (.sql)\n")
            f.write(f"- **Documentation:** 4 files (.md)\n")
            f.write(f"- **ETL Scripts:** 3 files (.sh, .sql, .py)\n")
            f.write(f"- **JSON Data:** 1 file (.json)\n\n")

            f.write("## Support\n\n")
            f.write("For questions about this analysis, contact your migration team lead.\n")

    def export_reports(self, output_dir):
        """Export all reports to files"""
        print("Exporting reports...")

        # Create main output directory
        os.makedirs(output_dir, exist_ok=True)

        # Create thematic subdirectories
        schema_dir = os.path.join(output_dir, SCHEMA_DIR)
        powerbi_dir = os.path.join(output_dir, POWERBI_DIR)
        quality_dir = os.path.join(output_dir, DATA_QUALITY_DIR)
        rel_dir = os.path.join(output_dir, RELATIONSHIPS_DIR)
        migration_dir = os.path.join(output_dir, MIGRATION_DIR)
        etl_dir = os.path.join(output_dir, ETL_DIR)
        analysis_dir = os.path.join(output_dir, ANALYSIS_DIR)

        for directory in [schema_dir, powerbi_dir, quality_dir, rel_dir, migration_dir, etl_dir, analysis_dir]:
            os.makedirs(directory, exist_ok=True)

        print(f"   Created thematic subdirectories")

        # ========================================
        # 07-ANALYSIS: Core Analysis Files
        # ========================================

        # Table summary Excel
        table_summary = []
        for t in self.report["table_details"]:
            table_summary.append({
                "access_name": t["name"],
                "postgresql_name": t["pg_name"],
                "columns": len(t["columns"]),
                "rows": t["row_count"],
                "primary_key": t.get("primary_key") or "NONE",
                "name_changed": t["name"] != t["pg_name"]
            })
        pd.DataFrame(table_summary).to_excel(f"{analysis_dir}/tables_summary.xlsx", index=False)
        print(f"   Saved: {ANALYSIS_DIR}/tables_summary.xlsx")

        # Columns detail Excel
        columns_detail = []
        for t in self.report["table_details"]:
            for c in t["columns"]:
                columns_detail.append({
                    "table": t["name"],
                    "column": c["name"],
                    "pg_column": c["pg_name"],
                    "type": c["type"],
                    "size": c["size"],
                    "nullable": c["nullable"],
                    "name_changed": c["name"] != c["pg_name"]
                })
        pd.DataFrame(columns_detail).to_excel(f"{analysis_dir}/columns_detail.xlsx", index=False)
        print(f"   Saved: {ANALYSIS_DIR}/columns_detail.xlsx")

        # ========================================
        # 05-MIGRATION-PLANNING: Planning & Execution Files
        # ========================================

        # Full JSON report
        with open(f"{migration_dir}/full_analysis.json", "w", encoding="utf-8") as f:
            json.dump(self.report, f, indent=2, default=str, ensure_ascii=False)
        print(f"   Saved: {MIGRATION_DIR}/full_analysis.json")

        # Migration Summary
        self.generate_summary(f"{migration_dir}/MIGRATION_SUMMARY.md")
        print(f"   Saved: {MIGRATION_DIR}/MIGRATION_SUMMARY.md")

        # Migration Checklist
        self.generate_migration_checklist_md(f"{migration_dir}/migration_checklist.md")
        print(f"   Saved: {MIGRATION_DIR}/migration_checklist.md")

        # Migration Review Checklist
        self.generate_migration_review_checklist(f"{migration_dir}/migration_review_checklist.xlsx")
        print(f"   Saved: {MIGRATION_DIR}/migration_review_checklist.xlsx")

        # ========================================
        # 04-RELATIONSHIPS: Foreign Keys & Relationships
        # ========================================

        # Relationships
        if self.report["relationships"]["details"]:
            pd.DataFrame(self.report["relationships"]["details"]).to_excel(
                f"{rel_dir}/relationships.xlsx", index=False
            )
            print(f"   Saved: {RELATIONSHIPS_DIR}/relationships.xlsx")

        # Inferred Foreign Keys
        if self.report.get("inferred_foreign_keys"):
            pd.DataFrame(self.report["inferred_foreign_keys"]).to_excel(
                f"{rel_dir}/inferred_foreign_keys.xlsx", index=False
            )
            print(f"   Saved: {RELATIONSHIPS_DIR}/inferred_foreign_keys.xlsx")

        # Indexes
        if self.report.get("indexes"):
            pd.DataFrame(self.report["indexes"]).to_excel(
                f"{rel_dir}/indexes.xlsx", index=False
            )
            print(f"   Saved: {RELATIONSHIPS_DIR}/indexes.xlsx")

        # ========================================
        # 03-DATA-QUALITY: Data Quality & Validation
        # ========================================

        # Data quality Excel
        quality_rows = []
        for tq in self.report["data_quality"]:
            for cq in tq["columns"]:
                quality_rows.append({
                    "table": tq["table"],
                    "column": cq["column"],
                    "null_count": cq.get("null_count"),
                    "null_percent": cq.get("null_percent"),
                    "distinct_count": cq.get("distinct_count"),
                    "sample_values": "; ".join(str(v) for v in cq.get("sample_values", []))
                })
        pd.DataFrame(quality_rows).to_excel(f"{quality_dir}/data_quality.xlsx", index=False)
        print(f"   Saved: {DATA_QUALITY_DIR}/data_quality.xlsx")

        # Issues Excel
        if self.report["potential_issues"]["details"]:
            pd.DataFrame(self.report["potential_issues"]["details"]).to_excel(
                f"{quality_dir}/issues.xlsx", index=False
            )
            print(f"   Saved: {DATA_QUALITY_DIR}/issues.xlsx")

        # Dead Columns Analysis
        if self.report.get("dead_columns"):
            pd.DataFrame(self.report["dead_columns"]).to_excel(
                f"{quality_dir}/dead_columns_analysis.xlsx", index=False
            )
            print(f"   Saved: {DATA_QUALITY_DIR}/dead_columns_analysis.xlsx")

        # Referential Integrity Issues
        if self.report.get("referential_integrity_issues"):
            pd.DataFrame(self.report["referential_integrity_issues"]).to_excel(
                f"{quality_dir}/referential_integrity_issues.xlsx", index=False
            )
            print(f"   Saved: {DATA_QUALITY_DIR}/referential_integrity_issues.xlsx")

        # Data Validation Queries SQL
        self.generate_data_validation_queries(f"{quality_dir}/data_validation_queries.sql")
        print(f"   Saved: {DATA_QUALITY_DIR}/data_validation_queries.sql")

        # ========================================
        # 02-POWERBI: Power BI Migration Files
        # ========================================

        # Power BI Impact Analysis
        if self.report.get("powerbi_impact"):
            pd.DataFrame(self.report["powerbi_impact"]).to_excel(
                f"{powerbi_dir}/powerbi_impact_analysis.xlsx", index=False
            )
            print(f"   Saved: {POWERBI_DIR}/powerbi_impact_analysis.xlsx")

        # DAX Impact Analysis
        if self.report.get("dax_impact"):
            pd.DataFrame(self.report["dax_impact"]).to_excel(
                f"{powerbi_dir}/dax_impact_analysis.xlsx", index=False
            )
            print(f"   Saved: {POWERBI_DIR}/dax_impact_analysis.xlsx")

        # Naming Mappings
        table_map, col_map = self.generate_naming_mapping()
        pd.DataFrame(table_map).to_excel(
            f"{powerbi_dir}/naming_mapping_tables.xlsx", index=False
        )
        print(f"   Saved: {POWERBI_DIR}/naming_mapping_tables.xlsx")

        if col_map:
            pd.DataFrame(col_map).to_excel(
                f"{powerbi_dir}/naming_mapping_columns.xlsx", index=False
            )
            print(f"   Saved: {POWERBI_DIR}/naming_mapping_columns.xlsx")

        # Connection Guide
        self.generate_connection_docs(powerbi_dir)
        print(f"   Saved: {POWERBI_DIR}/powerbi_connection_guide.md")

        # ========================================
        # 01-SCHEMA: Database Schema Files
        # ========================================

        # PostgreSQL schema
        self.generate_pg_schema(f"{schema_dir}/postgresql_schema.sql")
        print(f"   Saved: {SCHEMA_DIR}/postgresql_schema.sql")

        # Compatibility Views SQL
        self.generate_compatibility_views(f"{schema_dir}/postgresql_compatibility_views.sql")
        print(f"   Saved: {SCHEMA_DIR}/postgresql_compatibility_views.sql")

        # ========================================
        # 06-ETL: ETL Scripts & Data Transfer
        # ========================================

        # Queries list (Access saved queries)
        if self.report["queries"]["details"]:
            pd.DataFrame(self.report["queries"]["details"]).to_excel(
                f"{etl_dir}/queries.xlsx", index=False
            )
            print(f"   Saved: {ETL_DIR}/queries.xlsx")

        # ETL Scripts
        self.generate_etl_scripts(etl_dir)
        print(f"   Saved: {ETL_DIR}/01_export_from_access.sh")
        print(f"   Saved: {ETL_DIR}/02_import_to_postgres.sql")
        print(f"   Saved: {ETL_DIR}/03_transform_data.py")

        # ========================================
        # Generate README for output directory
        # ========================================
        self.generate_output_readme(output_dir)
        print(f"   Saved: README.md")

        print(f"\n{'='*60}")
        print("ORGANIZED OUTPUT STRUCTURE")
        print(f"{'='*60}")
        print(f"\nAll reports exported to: '{output_dir}/'")
        print(f"\n{SCHEMA_DIR}/ - Database Structure (2 files)")
        print(f"  └─ PostgreSQL schema & compatibility views")
        print(f"\n{POWERBI_DIR}/ - Power BI Migration (5 files)")
        print(f"  └─ Impact analysis, DAX changes, naming mappings, connection guide")
        print(f"\n{DATA_QUALITY_DIR}/ - Data Quality & Validation (5 files)")
        print(f"  └─ Quality metrics, issues, validation queries, dead columns")
        print(f"\n{RELATIONSHIPS_DIR}/ - Relationships & Keys (3 files)")
        print(f"  └─ Foreign keys, relationships, indexes")
        print(f"\n{MIGRATION_DIR}/ - Migration Planning (4 files)")
        print(f"  └─ Summary, checklists, review forms, full analysis JSON")
        print(f"\n{ETL_DIR}/ - ETL Scripts & Data (4 files + csv/)")
        print(f"  └─ Export/import scripts, transformations, queries, CSV data")
        print(f"\n{ANALYSIS_DIR}/ - Core Analysis Tables (2 files)")
        print(f"  └─ Tables & columns summaries")
        print(f"\n{'='*60}")
        print(f"TOTAL: ~26 files organized in 7 themed directories")
        print(f"{'='*60}")
    
    def generate_pg_schema(self, filepath):
        """Generate PostgreSQL schema - can also use mdb-schema directly"""
        type_map = {
            "COUNTER": "SERIAL",
            "LONG INTEGER": "INTEGER",
            "INTEGER": "INTEGER",
            "SMALLINT": "SMALLINT",
            "SINGLE": "REAL",
            "DOUBLE": "DOUBLE PRECISION",
            "CURRENCY": "NUMERIC(19,4)",
            "DATETIME": "TIMESTAMP",
            "BOOLEAN": "BOOLEAN",
            "TEXT": "TEXT",
            "MEMO": "TEXT",
            "VARCHAR": "VARCHAR",
            "LONGBINARY": "BYTEA",
            "OLE": "BYTEA",
            "BYTE": "SMALLINT",
        }
        
        with open(filepath, "w", encoding="utf-8") as f:
            f.write("-- PostgreSQL Schema Generated from Access Database\n")
            f.write(f"-- Generated: {datetime.now()}\n")
            f.write(f"-- Source: {self.db_path}\n\n")
            
            # Option 1: Use mdb-schema directly for PostgreSQL
            f.write("-- Auto-generated using mdb-schema:\n")
            try:
                schema = self.run_mdb_command("mdb-schema", "--postgres")
                f.write(schema)
            except:
                # Option 2: Manual generation
                for table in self.report["table_details"]:
                    f.write(f"\n-- Table: {table['name']}\n")
                    f.write(f'CREATE TABLE "{table["pg_name"]}" (\n')
                    
                    col_lines = []
                    for col in table["columns"]:
                        pg_type = type_map.get(col["type"].upper(), "TEXT")
                        if pg_type == "VARCHAR" and col["size"]:
                            pg_type = f"VARCHAR({col['size']})"
                        
                        nullable = "" if col["nullable"] else " NOT NULL"
                        col_lines.append(f'    "{col["pg_name"]}" {pg_type}{nullable}')
                    
                    f.write(",\n".join(col_lines))
                    f.write("\n);\n")
    
    def generate_migration_review_checklist(self, filepath):
        """Generate comprehensive migration review checklist for stakeholders"""
        print("   Generating migration review checklist...")

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        # Create Excel writer
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:

            # Sheet 1: Tables Review
            tables_review = []
            for table in self.report["table_details"]:
                # Count issues for this table
                table_issues = [i for i in self.report["potential_issues"]["details"]
                               if i.get("table") == table["name"]]

                # Detect potential deprecated tables
                is_deprecated = (
                    table["name"].lower().startswith("xxx") or
                    "paste error" in table["name"].lower() or
                    "copy of" in table["name"].lower() or
                    table["row_count"] == 0
                )

                # Estimate usage based on row count and name
                estimated_status = "UNKNOWN"
                if is_deprecated:
                    estimated_status = "DEPRECATED"
                elif table["row_count"] > 0:
                    estimated_status = "ACTIVE"
                else:
                    estimated_status = "EMPTY"

                # Suggest priority based on row count and issues
                suggested_priority = "MEDIUM"
                if table["row_count"] > 100:
                    suggested_priority = "HIGH"
                elif table["row_count"] > 20:
                    suggested_priority = "MEDIUM"
                else:
                    suggested_priority = "LOW"

                tables_review.append({
                    "Table_Name_Access": table["name"],
                    "Table_Name_PostgreSQL": table["pg_name"],
                    "Row_Count": table["row_count"],
                    "Column_Count": len(table["columns"]),
                    "Name_Will_Change": "YES" if table["name"] != table["pg_name"] else "NO",
                    "Estimated_Status": estimated_status,
                    "Suggested_Priority": suggested_priority,
                    "Issues_Count": len(table_issues),
                    "DECISION_Keep_or_Discard": "",  # Empty for user to fill
                    "DECISION_Final_Priority": "",  # Empty for user to fill
                    "DECISION_Notes": "",  # Empty for user to fill
                    "Used_in_PowerBI": ""  # Empty for user to fill
                })

            df_tables = pd.DataFrame(tables_review)
            df_tables.to_excel(writer, sheet_name='Tables_Review', index=False)

            # Sheet 2: Columns Review
            columns_review = []
            for table in self.report["table_details"]:
                for col in table["columns"]:
                    # Get quality info if available
                    quality_info = None
                    for tq in self.report.get("data_quality", []):
                        if tq["table"] == table["name"]:
                            for cq in tq["columns"]:
                                if cq["column"] == col["name"]:
                                    quality_info = cq
                                    break

                    # Detect potential issues
                    has_special_chars = any(c in col["name"] for c in [" ", "(", ")", "/", "%", "-"])

                    suggested_action = "KEEP"
                    if col["name"].lower().startswith("f") and col["name"][1:].isdigit():
                        # Generic column names like F1, F2, F10
                        suggested_action = "REVIEW"

                    columns_review.append({
                        "Table_Name": table["name"],
                        "Column_Name_Access": col["name"],
                        "Column_Name_PostgreSQL": col["pg_name"],
                        "Data_Type": col["type"],
                        "Size": col.get("size", ""),
                        "Nullable": "YES" if col["nullable"] else "NO",
                        "Name_Will_Change": "YES" if col["name"] != col["pg_name"] else "NO",
                        "Has_Special_Characters": "YES" if has_special_chars else "NO",
                        "Null_Percent": quality_info.get("null_percent", "") if quality_info else "",
                        "Distinct_Count": quality_info.get("distinct_count", "") if quality_info else "",
                        "Sample_Values": "; ".join(str(v)[:30] for v in quality_info.get("sample_values", [])[:3]) if quality_info else "",
                        "Suggested_Action": suggested_action,
                        "DECISION_Keep_or_Discard": "",  # Empty for user to fill
                        "DECISION_Notes": ""  # Empty for user to fill
                    })

            df_columns = pd.DataFrame(columns_review)
            df_columns.to_excel(writer, sheet_name='Columns_Review', index=False)

            # Sheet 3: Summary Dashboard
            summary_data = []

            # Table statistics
            total_tables = len(self.report["table_details"])
            deprecated_tables = sum(1 for t in tables_review if t["Estimated_Status"] == "DEPRECATED")
            active_tables = sum(1 for t in tables_review if t["Estimated_Status"] == "ACTIVE")
            empty_tables = sum(1 for t in tables_review if t["Estimated_Status"] == "EMPTY")

            summary_data.append({"Metric": "Total Tables", "Count": total_tables})
            summary_data.append({"Metric": "- Active Tables", "Count": active_tables})
            summary_data.append({"Metric": "- Deprecated/Test Tables", "Count": deprecated_tables})
            summary_data.append({"Metric": "- Empty Tables", "Count": empty_tables})
            summary_data.append({"Metric": "", "Count": ""})

            # Column statistics
            total_columns = sum(len(t["columns"]) for t in self.report["table_details"])
            columns_with_name_changes = sum(1 for c in columns_review if c["Name_Will_Change"] == "YES")
            columns_with_special_chars = sum(1 for c in columns_review if c["Has_Special_Characters"] == "YES")

            summary_data.append({"Metric": "Total Columns", "Count": total_columns})
            summary_data.append({"Metric": "- Columns with Name Changes", "Count": columns_with_name_changes})
            summary_data.append({"Metric": "- Columns with Special Characters", "Count": columns_with_special_chars})
            summary_data.append({"Metric": "", "Count": ""})

            # Data statistics
            total_rows = sum(t["row_count"] for t in self.report["table_details"] if isinstance(t["row_count"], int))
            summary_data.append({"Metric": "Total Data Rows", "Count": total_rows})
            summary_data.append({"Metric": "", "Count": ""})

            # Issues
            summary_data.append({"Metric": "Total Issues Identified", "Count": self.report["potential_issues"]["count"]})
            summary_data.append({"Metric": "- HIGH Severity", "Count": self.report["potential_issues"]["by_severity"]["HIGH"]})
            summary_data.append({"Metric": "- MEDIUM Severity", "Count": self.report["potential_issues"]["by_severity"]["MEDIUM"]})
            summary_data.append({"Metric": "- LOW Severity", "Count": self.report["potential_issues"]["by_severity"]["LOW"]})

            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary_Dashboard', index=False)

            # Sheet 4: Instructions
            instructions = [
                {"Section": "PURPOSE", "Instructions": "This checklist helps you review all database objects before migration"},
                {"Section": "", "Instructions": ""},
                {"Section": "HOW TO USE", "Instructions": ""},
                {"Section": "1. Tables Review", "Instructions": "Review each table and fill in the DECISION columns:"},
                {"Section": "", "Instructions": "  - DECISION_Keep_or_Discard: Enter KEEP, DISCARD, or REVIEW"},
                {"Section": "", "Instructions": "  - DECISION_Final_Priority: Enter CRITICAL, HIGH, MEDIUM, or LOW"},
                {"Section": "", "Instructions": "  - DECISION_Notes: Add any comments or reasons"},
                {"Section": "", "Instructions": "  - Used_in_PowerBI: Enter YES or NO if you know"},
                {"Section": "", "Instructions": ""},
                {"Section": "2. Columns Review", "Instructions": "Review columns for tables you're keeping:"},
                {"Section": "", "Instructions": "  - DECISION_Keep_or_Discard: Enter KEEP or DISCARD"},
                {"Section": "", "Instructions": "  - DECISION_Notes: Note why you're discarding (if applicable)"},
                {"Section": "", "Instructions": ""},
                {"Section": "3. Look for", "Instructions": ""},
                {"Section": "", "Instructions": "  - Tables with 'xxx' prefix (usually test/deprecated)"},
                {"Section": "", "Instructions": "  - Tables with 0 rows (may be obsolete)"},
                {"Section": "", "Instructions": "  - 'Copy Of' or 'Paste Error' tables (backups/errors)"},
                {"Section": "", "Instructions": "  - Columns with generic names (F1, F2, F10)"},
                {"Section": "", "Instructions": "  - Columns that are always empty (high null %)"},
                {"Section": "", "Instructions": ""},
                {"Section": "4. Save", "Instructions": "Save this file and share with migration team"},
                {"Section": "", "Instructions": ""},
                {"Section": "SUGGESTED PRIORITIES", "Instructions": ""},
                {"Section": "", "Instructions": "  - CRITICAL: Core business data, used in production reports"},
                {"Section": "", "Instructions": "  - HIGH: Important data, used regularly"},
                {"Section": "", "Instructions": "  - MEDIUM: Reference data, used occasionally"},
                {"Section": "", "Instructions": "  - LOW: Archive data, rarely used"},
                {"Section": "", "Instructions": ""},
                {"Section": "QUESTIONS?", "Instructions": "Contact your migration team lead"}
            ]

            df_instructions = pd.DataFrame(instructions)
            df_instructions.to_excel(writer, sheet_name='Instructions', index=False)

            # ========================================
            # FORMATTING SECTION
            # ========================================

            # Get workbook to apply formatting
            workbook = writer.book

            # --- Format Sheet 1: Tables_Review ---
            ws_tables = workbook['Tables_Review']

            # Define styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            decision_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            decision_font = Font(bold=True, color="C65911", size=10)
            center_alignment = Alignment(horizontal="center", vertical="center")
            wrap_alignment = Alignment(wrap_text=True, vertical="top")

            # Format headers
            for cell in ws_tables[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            # Highlight DECISION columns with yellow background
            decision_columns_tables = [9, 10, 11, 12]  # DECISION_Keep_or_Discard, DECISION_Final_Priority, DECISION_Notes, Used_in_PowerBI
            for row in range(2, ws_tables.max_row + 1):
                for col_idx in decision_columns_tables:
                    cell = ws_tables.cell(row=row, column=col_idx)
                    cell.fill = decision_fill
                    if col_idx in [9, 10, 12]:  # Not the notes column
                        cell.font = decision_font
                        cell.alignment = center_alignment

            # Set column widths
            column_widths_tables = {
                'A': 35,  # Table_Name_Access
                'B': 35,  # Table_Name_PostgreSQL
                'C': 12,  # Row_Count
                'D': 12,  # Column_Count
                'E': 15,  # Name_Will_Change
                'F': 15,  # Estimated_Status
                'G': 18,  # Suggested_Priority
                'H': 12,  # Issues_Count
                'I': 22,  # DECISION_Keep_or_Discard
                'J': 22,  # DECISION_Final_Priority
                'K': 40,  # DECISION_Notes
                'L': 18,  # Used_in_PowerBI
            }
            for col, width in column_widths_tables.items():
                ws_tables.column_dimensions[col].width = width

            # Add data validation (dropdowns) for decision columns
            # DECISION_Keep_or_Discard dropdown
            dv_keep_discard = DataValidation(type="list", formula1='"KEEP,DISCARD,REVIEW"', allow_blank=True)
            dv_keep_discard.error = 'Please select KEEP, DISCARD, or REVIEW'
            dv_keep_discard.errorTitle = 'Invalid Entry'
            ws_tables.add_data_validation(dv_keep_discard)
            dv_keep_discard.add(f'I2:I{ws_tables.max_row}')

            # DECISION_Final_Priority dropdown
            dv_priority = DataValidation(type="list", formula1='"CRITICAL,HIGH,MEDIUM,LOW"', allow_blank=True)
            dv_priority.error = 'Please select CRITICAL, HIGH, MEDIUM, or LOW'
            dv_priority.errorTitle = 'Invalid Entry'
            ws_tables.add_data_validation(dv_priority)
            dv_priority.add(f'J2:J{ws_tables.max_row}')

            # Used_in_PowerBI dropdown
            dv_powerbi = DataValidation(type="list", formula1='"YES,NO,UNKNOWN"', allow_blank=True)
            dv_powerbi.error = 'Please select YES, NO, or UNKNOWN'
            dv_powerbi.errorTitle = 'Invalid Entry'
            ws_tables.add_data_validation(dv_powerbi)
            dv_powerbi.add(f'L2:L{ws_tables.max_row}')

            # Freeze top row
            ws_tables.freeze_panes = 'A2'

            # --- Format Sheet 2: Columns_Review ---
            ws_columns = workbook['Columns_Review']

            # Format headers
            for cell in ws_columns[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            # Highlight DECISION columns (columns 12 and 13)
            decision_columns_cols = [12, 13]  # DECISION_Keep_or_Discard, DECISION_Notes
            for row in range(2, ws_columns.max_row + 1):
                for col_idx in decision_columns_cols:
                    cell = ws_columns.cell(row=row, column=col_idx)
                    cell.fill = decision_fill
                    if col_idx == 12:
                        cell.font = decision_font
                        cell.alignment = center_alignment

            # Set column widths
            column_widths_cols = {
                'A': 30,  # Table_Name
                'B': 30,  # Column_Name_Access
                'C': 30,  # Column_Name_PostgreSQL
                'D': 15,  # Data_Type
                'E': 8,   # Size
                'F': 10,  # Nullable
                'G': 15,  # Name_Will_Change
                'H': 20,  # Has_Special_Characters
                'I': 12,  # Null_Percent
                'J': 12,  # Distinct_Count
                'K': 35,  # Sample_Values
                'L': 15,  # Suggested_Action
                'M': 22,  # DECISION_Keep_or_Discard
                'N': 40,  # DECISION_Notes
            }
            for col, width in column_widths_cols.items():
                ws_columns.column_dimensions[col].width = width

            # Add data validation for DECISION_Keep_or_Discard
            dv_col_decision = DataValidation(type="list", formula1='"KEEP,DISCARD,REVIEW"', allow_blank=True)
            dv_col_decision.error = 'Please select KEEP, DISCARD, or REVIEW'
            dv_col_decision.errorTitle = 'Invalid Entry'
            ws_columns.add_data_validation(dv_col_decision)
            dv_col_decision.add(f'M2:M{ws_columns.max_row}')

            # Freeze top row
            ws_columns.freeze_panes = 'A2'

            # --- Format Sheet 3: Summary_Dashboard ---
            ws_summary = workbook['Summary_Dashboard']

            # Format headers
            for cell in ws_summary[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            # Bold and larger font for summary
            for row in ws_summary.iter_rows(min_row=2):
                row[0].font = Font(bold=True, size=10)
                row[1].font = Font(size=10)

            # Set column widths
            ws_summary.column_dimensions['A'].width = 35
            ws_summary.column_dimensions['B'].width = 15

            # --- Format Sheet 4: Instructions ---
            ws_instructions = workbook['Instructions']

            # Format headers
            for cell in ws_instructions[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            # Wrap text for instructions
            for row in ws_instructions.iter_rows(min_row=2):
                row[0].font = Font(bold=True)
                row[1].alignment = wrap_alignment

            # Set column widths
            ws_instructions.column_dimensions['A'].width = 20
            ws_instructions.column_dimensions['B'].width = 80

        print(f"      Created {len(tables_review)} table reviews")
        print(f"      Created {len(columns_review)} column reviews")
        print(f"      Applied formatting with dropdowns and color coding")

    def generate_summary(self, filepath):
        """Generate executive summary"""
        with open(filepath, "w", encoding="utf-8") as f:
            f.write("# Access to PostgreSQL Migration Summary\n\n")
            f.write(f"**Source Database:** `{self.db_path}`\n\n")
            f.write(f"**Analysis Date:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n")

            f.write("## Overview\n\n")
            f.write("| Metric | Count |\n")
            f.write("|--------|-------|\n")
            f.write(f"| Tables | {self.report['tables']['count']} |\n")
            f.write(f"| Saved Queries | {self.report['queries']['count']} |\n")
            f.write(f"| Relationships | {self.report['relationships']['count']} |\n")

            total_rows = sum(t['row_count'] for t in self.report['table_details'] if isinstance(t['row_count'], int))
            f.write(f"| Total Rows | {total_rows:,} |\n\n")

            f.write("## Tables\n\n")
            f.write("| Access Table | PostgreSQL Table | Rows | Columns |\n")
            f.write("|--------------|------------------|------|--------|\n")
            for t in self.report["table_details"]:
                row_count = t['row_count'] if isinstance(t['row_count'], int) else 0
                f.write(f"| {t['name']} | {t['pg_name']} | {row_count:,} | {len(t['columns'])} |\n")

            f.write("\n## Potential Issues\n\n")
            issues = self.report["potential_issues"]
            f.write(f"**Total:** {issues['count']} ")
            f.write(f"(HIGH: {issues['by_severity']['HIGH']}, ")
            f.write(f"MEDIUM: {issues['by_severity']['MEDIUM']}, ")
            f.write(f"LOW: {issues['by_severity']['LOW']})\n\n")

            high_issues = [i for i in issues["details"] if i["severity"] == "HIGH"]
            if high_issues:
                f.write("### High Priority\n\n")
                for issue in high_issues:
                    f.write(f"- **{issue['type']}**: {issue['issue']}\n")


if __name__ == "__main__":
    analyzer = AccessDatabaseAnalyzerWSL(ACCESS_PATH)
    
    try:
        analyzer.analyze_all()
        analyzer.export_reports(OUTPUT_DIR)
    finally:
        pass
    
    print("\n" + "=" * 60)
    print("ANALYSIS COMPLETE")
    print("=" * 60)
    print(f"\nReview the reports in '{OUTPUT_DIR}/' folder")